from __future__ import annotations

import difflib
import json
import re
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.services.calendar import import_calendar_saved_notice
from app.services.notice_meta import classify_notice_tag
from app.types import NoticeCandidate
from app.utils import extract_datetimes, normalize_text

DATE_RANGE_RE = re.compile(
    r"(?P<start>\d{4}[./-]\d{2}[./-]\d{2})(?:\s*[~\-]\s*)(?P<end>\d{4}[./-]\d{2}[./-]\d{2})"
)


@dataclass
class ImportedHistoryRow:
    year: int | None
    organization: str
    title: str
    contract_no: str
    period: str
    amount_value: int | None
    amount_text: str | None
    primary_deadline_at: datetime | None
    deadline_confidence: str
    notice_tag: str
    raw_payload: dict


@dataclass
class HistorySearchMatch:
    row: ImportedHistoryRow
    site_code: str
    site_name: str
    search_term: str
    candidate: NoticeCandidate
    score: float


def _parse_amount(value) -> tuple[int | None, str | None]:
    if value is None or value == "":
        return None, None
    if isinstance(value, (int, float)):
        amount_value = int(value)
        return amount_value, f"{amount_value:,}원"
    text = normalize_text(str(value))
    digits = re.sub(r"[^\d]", "", text)
    if not digits:
        return None, text or None
    amount_value = int(digits)
    return amount_value, text


def _estimate_deadline(period: str) -> tuple[datetime | None, str]:
    period_text = normalize_text(period)
    if not period_text:
        return None, "unknown"

    match = DATE_RANGE_RE.search(period_text)
    if match:
        start_raw = match.group("start").replace(".", "-").replace("/", "-")
        try:
            start_dt = datetime.strptime(start_raw, "%Y-%m-%d").replace(
                hour=18, minute=0, second=0, microsecond=0
            )
            return start_dt, "estimated"
        except ValueError:
            return None, "unknown"

    explicit_markers = ("\uc785\ucc30", "\uc81c\ucd9c", "\ub9c8\uac10", "\uc81c\uc548")
    if any(marker in period_text for marker in explicit_markers):
        compact_dates = re.findall(r"\d{4}[./-]\d{2}[./-]\d{2}", period_text)
        if compact_dates:
            raw = compact_dates[0].replace(".", "-").replace("/", "-")
            try:
                return (
                    datetime.strptime(raw, "%Y-%m-%d").replace(
                        hour=18, minute=0, second=0, microsecond=0
                    ),
                    "exact",
                )
            except ValueError:
                return None, "unknown"

    return None, "unknown"


def _classify_notice_tag(title: str, organization: str, period: str) -> str:
    candidate = NoticeCandidate(
        site_code="imported",
        site_notice_key=title,
        title=title,
        source_url="",
        organization=organization,
        period_text=period,
        raw_payload={},
    )
    return classify_notice_tag(candidate)


def _compact_text(value: str) -> str:
    text = normalize_text(value).lower()
    text = re.sub(r"^\d{4}\s*년\s*", "", text)
    text = re.sub(r"\([^)]*\)", " ", text)
    text = re.sub(r"[^0-9a-z가-힣]+", "", text)
    return text


def _word_terms(title: str) -> list[str]:
    normalized = normalize_text(title)
    stripped = re.sub(r"^\d{4}\s*년\s*", "", normalized)
    stripped = re.sub(r"\([^)]*\)", " ", stripped)
    stripped = normalize_text(stripped)
    words = [word for word in stripped.split(" ") if len(word) >= 2]
    terms: list[str] = []
    if stripped:
        terms.append(stripped)
    if len(words) >= 3:
        terms.append(" ".join(words[:3]))
    if len(words) >= 2:
        terms.append(" ".join(words[:2]))
    if words:
        terms.append(words[0])
    seen = set()
    deduped = []
    for term in terms:
        key = normalize_text(term)
        if key and key not in seen:
            seen.add(key)
            deduped.append(key)
    return deduped[:4]


def _similarity(left: str, right: str) -> float:
    return difflib.SequenceMatcher(None, _compact_text(left), _compact_text(right)).ratio()


def _organization_bonus(row_org: str, candidate_org: str) -> float:
    left = _compact_text(row_org)
    right = _compact_text(candidate_org)
    if not left or not right:
        return 0.0
    if left in right or right in left:
        return 0.15
    return 0.0


def _score_candidate(row: ImportedHistoryRow, candidate: NoticeCandidate) -> float:
    candidate_deadline = _candidate_deadline(candidate)
    if row.year is not None and candidate_deadline is not None and candidate_deadline.year != row.year:
        return -1.0
    title_score = _similarity(row.title, candidate.title)
    org_bonus = _organization_bonus(row.organization, candidate.organization or "")
    deadline_bonus = 0.05 if candidate_deadline is not None else 0.0
    return title_score + org_bonus + deadline_bonus


def _candidate_deadline(candidate: NoticeCandidate) -> datetime | None:
    if candidate.deadline_at is not None:
        return candidate.deadline_at
    dates = extract_datetimes(candidate.period_text or "")
    if len(dates) >= 2:
        return dates[-1]
    if dates:
        return dates[0]
    return None


def _find_best_match(
    row: ImportedHistoryRow,
    collector_registry: dict,
    *,
    site_codes: Iterable[str] = ("g2b", "iris"),
) -> HistorySearchMatch | None:
    best: HistorySearchMatch | None = None
    for site_code in site_codes:
        collector = collector_registry.get(site_code)
        if collector is None:
            continue
        for term in _word_terms(row.title):
            try:
                candidates = collector.search(term)
            except Exception:
                continue
            for candidate in candidates[:10]:
                score = _score_candidate(row, candidate)
                if best is None or score > best.score:
                    site_name = {
                        "g2b": "나라장터",
                        "iris": "IRIS",
                    }.get(site_code, site_code.upper())
                    best = HistorySearchMatch(
                        row=row,
                        site_code=site_code,
                        site_name=site_name,
                        search_term=term,
                        candidate=candidate,
                        score=score,
                    )
    if best is None or best.score < 0.52:
        return None
    return best


def _filter_rows_by_years(
    rows: list[ImportedHistoryRow],
    target_years: Iterable[int] | None,
) -> list[ImportedHistoryRow]:
    if not target_years:
        return rows
    year_set = {int(year) for year in target_years}
    return [row for row in rows if row.year in year_set]


def _iter_excel_rows(input_path: Path) -> list[ImportedHistoryRow]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    try:
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []

        current_year: int | None = None
        imported_rows: list[ImportedHistoryRow] = []
        for row in rows[1:]:
            values = list(row or ())
            while len(values) < 6:
                values.append(None)

            raw_year, raw_org, raw_title, raw_contract_no, raw_period, raw_amount = values[:6]

            if raw_year not in (None, ""):
                try:
                    current_year = int(raw_year)
                except (TypeError, ValueError):
                    year_text = normalize_text(str(raw_year))
                    digits = re.sub(r"[^\d]", "", year_text)
                    current_year = int(digits[:4]) if len(digits) >= 4 else current_year

            title = normalize_text(str(raw_title or ""))
            if not title:
                continue

            organization = normalize_text(str(raw_org or ""))
            contract_no = normalize_text(str(raw_contract_no or ""))
            period = normalize_text(str(raw_period or ""))
            amount_value, amount_text = _parse_amount(raw_amount)
            primary_deadline_at, deadline_confidence = _estimate_deadline(period)
            notice_tag = _classify_notice_tag(title, organization, period)

            imported_rows.append(
                ImportedHistoryRow(
                    year=current_year,
                    organization=organization,
                    title=title,
                    contract_no=contract_no,
                    period=period,
                    amount_value=amount_value,
                    amount_text=amount_text,
                    primary_deadline_at=primary_deadline_at,
                    deadline_confidence=deadline_confidence,
                    notice_tag=notice_tag,
                    raw_payload={
                        "year": current_year,
                        "organization": organization,
                        "title": title,
                        "contract_no": contract_no,
                        "period": period,
                        "amount_text": amount_text,
                    },
                )
            )

        return imported_rows
    finally:
        workbook.close()


def preview_history_import(input_path: Path, now: datetime) -> dict:
    rows = _iter_excel_rows(input_path)
    displayable_start_year = now.year - 3
    displayable_rows = [
        row
        for row in rows
        if row.year is not None and row.year >= displayable_start_year and row.year <= now.year
    ]
    calendar_displayable = [
        row
        for row in displayable_rows
        if row.primary_deadline_at is not None and row.deadline_confidence != "unknown"
    ]

    report = {
        "input_file": str(input_path),
        "total_rows": len(rows),
        "within_calendar_range": len(displayable_rows),
        "calendar_displayable": len(calendar_displayable),
        "exact": sum(1 for row in rows if row.deadline_confidence == "exact"),
        "estimated": sum(1 for row in rows if row.deadline_confidence == "estimated"),
        "unknown": sum(1 for row in rows if row.deadline_confidence == "unknown"),
        "years": {},
        "preview_items": [],
    }

    year_counts: dict[str, int] = {}
    for row in rows:
        label = str(row.year or "unknown")
        year_counts[label] = year_counts.get(label, 0) + 1
    report["years"] = year_counts

    for row in calendar_displayable[:15]:
        report["preview_items"].append(
            {
                "year": row.year,
                "title": row.title,
                "organization": row.organization,
                "primary_deadline_at": (
                    row.primary_deadline_at.isoformat() if row.primary_deadline_at else None
                ),
                "deadline_confidence": row.deadline_confidence,
                "amount_text": row.amount_text,
                "notice_tag": row.notice_tag,
            }
        )

    return report


def apply_history_import(
    session: Session,
    input_path: Path,
    *,
    now: datetime,
    selected_by: str,
    import_batch_id: str,
) -> dict:
    rows = _iter_excel_rows(input_path)
    displayable_start_year = now.year - 3
    imported = 0
    skipped = 0

    for row in rows:
        if row.year is None or row.year < displayable_start_year or row.year > now.year:
            skipped += 1
            continue
        if row.primary_deadline_at is None:
            skipped += 1
            continue

        import_calendar_saved_notice(
            session=session,
            title=row.title,
            organization=row.organization or None,
            site_name="과거이관",
            primary_deadline_at=row.primary_deadline_at,
            amount_text_value=row.amount_text,
            amount_value=row.amount_value,
            priority_score=3,
            notice_tag=row.notice_tag,
            source_url="",
            status="closed",
            owner_name=None,
            memo=None,
            selected_by=selected_by,
            deadline_confidence=row.deadline_confidence,
            raw_payload={
                **row.raw_payload,
                "origin_type": "imported",
                "import_batch_id": import_batch_id,
            },
            legacy_year=row.year,
            import_batch_id=import_batch_id,
        )
        imported += 1

    return {
        "input_file": str(input_path),
        "import_batch_id": import_batch_id,
        "imported": imported,
        "skipped": skipped,
        "total_rows": len(rows),
    }


def preview_history_site_search_import(
    input_path: Path,
    *,
    now: datetime,
    collector_registry: dict,
    target_years: Iterable[int] | None = None,
    site_codes: Iterable[str] = ("g2b", "iris"),
) -> dict:
    rows = _iter_excel_rows(input_path)
    default_target_rows = [row for row in rows if row.year in (2023, 2024, 2025)]
    target_rows = _filter_rows_by_years(default_target_rows, target_years)
    matched: list[dict] = []
    unmatched: list[dict] = []

    for row in target_rows:
        match = _find_best_match(row, collector_registry, site_codes=site_codes)
        if match is None:
            unmatched.append(
                {
                    "year": row.year,
                    "title": row.title,
                    "organization": row.organization,
                    "deadline_confidence": row.deadline_confidence,
                }
            )
            continue
        matched.append(
            {
                "year": row.year,
                "title": row.title,
                "organization": row.organization,
                "site_code": match.site_code,
                "site_name": match.site_name,
                "search_term": match.search_term,
                "matched_title": match.candidate.title,
                "matched_organization": match.candidate.organization,
                "matched_deadline_at": (
                    _candidate_deadline(match.candidate).isoformat()
                    if _candidate_deadline(match.candidate)
                    else None
                ),
                "score": round(match.score, 4),
                "source_url": match.candidate.source_url,
            }
        )

    return {
        "input_file": str(input_path),
        "target_years": sorted({row.year for row in target_rows if row.year is not None}),
        "site_codes": list(site_codes),
        "target_rows": len(target_rows),
        "matched_rows": len(matched),
        "unmatched_rows": len(unmatched),
        "matched": matched[:50],
        "unmatched_preview": unmatched[:50],
    }


def apply_history_site_search_import(
    session: Session,
    input_path: Path,
    *,
    now: datetime,
    selected_by: str,
    import_batch_id: str,
    collector_registry: dict,
    target_years: Iterable[int] | None = None,
    site_codes: Iterable[str] = ("g2b", "iris"),
) -> dict:
    rows = _iter_excel_rows(input_path)
    default_target_rows = [row for row in rows if row.year in (2023, 2024, 2025)]
    target_rows = _filter_rows_by_years(default_target_rows, target_years)
    matched_count = 0
    fallback_count = 0
    skipped_count = 0

    for row in target_rows:
        match = _find_best_match(row, collector_registry, site_codes=site_codes)
        if match is not None:
            deadline_at = _candidate_deadline(match.candidate) or row.primary_deadline_at
            import_calendar_saved_notice(
                session=session,
                title=row.title,
                organization=row.organization or match.candidate.organization or None,
                site_code=match.site_code,
                site_name=match.site_name,
                primary_deadline_at=deadline_at,
                amount_text_value=row.amount_text,
                amount_value=row.amount_value,
                priority_score=3,
                notice_tag=row.notice_tag,
                source_url=match.candidate.source_url,
                status="closed",
                owner_name=None,
                memo=None,
                selected_by=selected_by,
                deadline_confidence="exact" if deadline_at else row.deadline_confidence,
                raw_payload={
                    **row.raw_payload,
                    "origin_type": "imported",
                    "import_batch_id": import_batch_id,
                    "matched_site_code": match.site_code,
                    "matched_site_name": match.site_name,
                    "search_term": match.search_term,
                    "matched_title": match.candidate.title,
                    "matched_organization": match.candidate.organization,
                    "matched_source_url": match.candidate.source_url,
                    "match_score": match.score,
                },
                legacy_year=row.year,
                import_batch_id=import_batch_id,
            )
            matched_count += 1
            continue

        if row.primary_deadline_at is None:
            skipped_count += 1
            continue

        import_calendar_saved_notice(
            session=session,
            title=row.title,
            organization=row.organization or None,
            site_code="imported",
            site_name="과거이관",
            primary_deadline_at=row.primary_deadline_at,
            amount_text_value=row.amount_text,
            amount_value=row.amount_value,
            priority_score=3,
            notice_tag=row.notice_tag,
            source_url="",
            status="closed",
            owner_name=None,
            memo=None,
            selected_by=selected_by,
            deadline_confidence=row.deadline_confidence,
            raw_payload={
                **row.raw_payload,
                "origin_type": "imported",
                "import_batch_id": import_batch_id,
                "matched_site_code": None,
                "matched_site_name": None,
                "search_term": None,
                "matched_title": None,
                "matched_organization": None,
                "matched_source_url": None,
                "match_score": None,
            },
            legacy_year=row.year,
            import_batch_id=import_batch_id,
        )
        fallback_count += 1

    return {
        "input_file": str(input_path),
        "import_batch_id": import_batch_id,
        "target_years": sorted({row.year for row in target_rows if row.year is not None}),
        "site_codes": list(site_codes),
        "target_rows": len(target_rows),
        "matched_imported": matched_count,
        "fallback_imported": fallback_count,
        "skipped": skipped_count,
    }


def write_preview_report(report: dict, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
