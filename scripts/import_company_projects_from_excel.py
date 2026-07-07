from __future__ import annotations

import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


TOKEN_RE = re.compile(r"[\uAC00-\uD7A3A-Za-z0-9]+")
ROMAN_RE = re.compile(r"[IVX]+")
STOPWORDS = {
    "사업",
    "용역",
    "연구",
    "개발",
    "구축",
    "개선",
    "운영",
    "관리",
    "지원",
    "기술",
    "서비스",
    "시스템",
    "정보",
    "자료",
    "분석",
    "활용",
    "기상",
    "해양",
    "수산",
    "모델",
    "예보",
    "모니터링",
    "계획",
    "추진",
    "안내",
    "공고",
    "교육",
    "센터",
    "기반",
    "고도화",
    "강화",
    "플랫폼",
    "유지",
    "보수",
    "및",
    "위한",
    "관련",
    "대한",
    "맞춤형",
    "국제",
    "국가",
    "한국",
    "사업화",
    "프로그램",
}


def normalize_text(value: str) -> str:
    return re.sub(r"\s+", " ", (value or "").replace("\r", " ").replace("\n", " ")).strip()


def read_excel_rows(input_path: Path) -> list[dict]:
    workbook = load_workbook(input_path, read_only=True, data_only=True)
    worksheet = workbook.active
    rows = worksheet.iter_rows(values_only=True)
    next(rows, None)  # header

    result = []
    current_year = ""
    for row in rows:
        values = list(row or ())
        while len(values) < 6:
            values.append("")

        year = normalize_text(str(values[0] or "")) or current_year
        current_year = year or current_year
        organization = normalize_text(str(values[1] or ""))
        project_name = normalize_text(str(values[2] or ""))
        contract_no = normalize_text(str(values[3] or ""))
        period = normalize_text(str(values[4] or ""))
        amount = normalize_text(str(values[5] or ""))

        result.append(
            {
                "year": year,
                "organization": organization,
                "project_name": project_name,
                "contract_no": contract_no,
                "period": period,
                "amount": amount,
            }
        )
    return result


def make_keywords(project_name: str) -> list[str]:
    text = normalize_text(project_name)
    tokens = TOKEN_RE.findall(text)

    filtered = []
    for token in tokens:
        if token.isdigit():
            continue
        if len(token) < 2 and not token.isupper():
            continue
        if token in STOPWORDS:
            continue
        if ROMAN_RE.fullmatch(token):
            continue
        filtered.append(token)

    unique_tokens = []
    for token in filtered:
        if token not in unique_tokens:
            unique_tokens.append(token)

    bigrams = []
    for left, right in zip(unique_tokens, unique_tokens[1:]):
        phrase = f"{left} {right}"
        if phrase not in bigrams:
            bigrams.append(phrase)

    keywords = []
    if text:
        keywords.append(text)
    keywords.extend(unique_tokens[:8])
    keywords.extend(bigrams[:6])

    final = []
    for keyword in keywords:
        keyword = normalize_text(keyword)
        if keyword and keyword not in final:
            final.append(keyword)
    return final


def transform_rows(raw_rows: list[dict]) -> list[dict]:
    projects = []
    seen = set()

    for row in raw_rows:
        year = normalize_text(row.get("year", ""))
        organization = normalize_text(row.get("organization", ""))
        project_name = normalize_text(row.get("project_name", ""))
        if not project_name:
            continue

        dedupe_key = (year, organization, project_name)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        projects.append(
            {
                "project_name": project_name,
                "organization": organization,
                "category": year or "수주이력",
                "keywords": make_keywords(project_name),
                "enabled": True,
            }
        )

    return projects


def main() -> None:
    parser = argparse.ArgumentParser(description="Convert awarded-project Excel to seed JSON")
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args()

    input_path = Path(args.input).expanduser().resolve()
    output_path = Path(args.output).expanduser().resolve()
    output_path.parent.mkdir(parents=True, exist_ok=True)

    raw_rows = read_excel_rows(input_path)
    projects = transform_rows(raw_rows)
    output_path.write_text(
        json.dumps(projects, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"input={input_path}")
    print(f"output={output_path}")
    print(f"count={len(projects)}")
    print(json.dumps(projects[:3], ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
